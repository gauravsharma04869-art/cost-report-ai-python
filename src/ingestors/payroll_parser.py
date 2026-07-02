"""
Payroll Register Parser — extracts salary and wage data broken down
by department, cost center, and employee type.

Handles:
- Excel payroll summaries
- CSV payroll exports from HR systems
- PS&R-style text summaries of salary allocations
"""

from __future__ import annotations

import csv
import io
import re
import time
from pathlib import Path
from typing import Any

from src.core.models import EntryType, GrossGLTransaction, ParsedIngestionResult
from src.ingestors.base_parser import BaseParser, detect_column_mapping, sanitize_amount, sanitize_string


class PayrollParser(BaseParser):
    """Parser for payroll register / salary allocation reports."""

    entry_type = EntryType.PAYROLL_REGISTER

    # Keywords that indicate payroll-related columns
    SALARY_KEYWORDS = [
        "salary", "wage", "wages", "payroll", "compensation", "labor",
        "hourly", "regular", "overtime", "bonus", "shift diff",
    ]

    def parse(self, content: bytes, source_file: str, **kwargs) -> ParsedIngestionResult:
        start_time = time.time()
        ext = Path(source_file).suffix.lower()
        errors: list[dict] = []

        try:
            if ext in (".xlsx", ".xls"):
                rows, col_mapping = self._parse_excel(content)
            else:
                rows, col_mapping = self._parse_csv(content)
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return ParsedIngestionResult(
                source_file=source_file,
                entry_type=self.entry_type,
                entries=[],
                row_count=0,
                error_count=1,
                errors=[{"row": 0, "error": f"Payroll parse failure: {e}"}],
                column_mapping={},
                processing_time_ms=round(elapsed, 2),
            )

        entries: list[GrossGLTransaction] = []
        for row in rows:
            try:
                # Determine account number: use department code or employee category
                acct_no = sanitize_string(row.get("account_number", ""))
                if not acct_no:
                    dept = sanitize_string(row.get("department_code", row.get("dept", "")))
                    emp_type = sanitize_string(row.get("employee_type", row.get("category", "")))
                    acct_no = f"{dept}-{emp_type}" if dept and emp_type else dept or "PAYROLL"

                entries.append(GrossGLTransaction(
                    account_number=acct_no,
                    account_description=sanitize_string(
                        row.get("account_description",
                               row.get("description",
                                      f"Payroll - {row.get('department_name', row.get('dept_name', ''))}"))
                    ),
                    debit_amount=sanitize_amount(
                        row.get("debit_amount", row.get("salary", row.get("total", "0")))
                    ),
                    credit_amount=sanitize_amount(row.get("credit_amount", "0")),
                    period=sanitize_string(row.get("period", row.get("pay_period", ""))),
                    department_code=sanitize_string(row.get("department_code", row.get("dept", ""))),
                    department_name=sanitize_string(row.get("department_name", row.get("dept_name", ""))),
                    source_file=source_file,
                    row_index=row.get("_row_index"),
                ))
            except Exception as e:
                errors.append({"row": row.get("_row_index"), "error": str(e)})

        elapsed = (time.time() - start_time) * 1000
        return ParsedIngestionResult(
            source_file=source_file,
            entry_type=self.entry_type,
            entries=entries,
            row_count=len(entries),
            error_count=len(errors),
            errors=errors,
            column_mapping=col_mapping,
            processing_time_ms=round(elapsed, 2),
        )

    def _parse_excel(self, content: bytes) -> tuple[list[dict], dict[str, str]]:
        import openpyxl

        rows: list[dict] = []
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        if sheet.max_row and sheet.max_row >= 2:
            headers = [str(cell.value or "") for cell in sheet[1]]
            col_mapping = detect_column_mapping(headers)

            # Payroll-specific header detection for salary columns
            salary_col = None
            for h in headers:
                h_lower = h.strip().lower()
                if any(kw in h_lower for kw in self.SALARY_KEYWORDS):
                    salary_col = headers.index(h)
                    break

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict: dict = {"_row_index": row_idx}
                for col_idx, header in enumerate(headers):
                    val = row[col_idx] if col_idx < len(row) else None
                    key = header.strip().lower().replace(" ", "_").replace("#", "")
                    if salary_col is not None and col_idx == salary_col:
                        row_dict["salary"] = val
                    elif key in ("total", "amount", "wages"):
                        row_dict["salary"] = val
                    row_dict[key] = val
                rows.append(row_dict)

        workbook.close()
        return rows, {}

    def _parse_csv(self, content: bytes) -> tuple[list[dict], dict[str, str]]:
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows: list[dict] = []

        try:
            headers = next(reader)
        except StopIteration:
            return rows, {}

        headers = [h.strip() for h in headers]
        col_mapping = detect_column_mapping(headers)

        for row_idx, row in enumerate(reader, start=2):
            row_dict: dict = {"_row_index": row_idx}
            for col_idx, val in enumerate(row):
                if col_idx < len(headers):
                    row_dict[headers[col_idx].strip().lower().replace(" ", "_")] = val.strip()
            rows.append(row_dict)

        return rows, col_mapping
