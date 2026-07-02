"""
Census Report Parser — extracts patient day, visit, or encounter counts
from EMR-generated PDF and Excel reports.

Handles:
- PDF census summaries (pdfplumber-based extraction)
- Excel census reports with date/department/day columns
- Structured text tables
"""

from __future__ import annotations

import io
import re
import time
from pathlib import Path
from typing import Any, Optional

from src.core.models import EntryType, GrossGLTransaction, ParsedIngestionResult
from src.ingestors.base_parser import BaseParser, detect_column_mapping, sanitize_amount, sanitize_string


class CensusParser(BaseParser):
    """Parser for census reports — patient days, visits, encounters."""

    entry_type = EntryType.CENSUS_REPORT

    def parse(self, content: bytes, source_file: str, **kwargs) -> ParsedIngestionResult:
        start_time = time.time()
        ext = Path(source_file).suffix.lower()
        errors: list[dict] = []

        try:
            if ext == ".pdf":
                rows, col_mapping = self._parse_pdf(content)
            elif ext in (".xlsx", ".xls"):
                rows, col_mapping = self._parse_excel(content, ext)
            elif ext in (".csv", ".tsv", ".txt"):
                rows, col_mapping = self._parse_csv_or_tsv(content, ext)
            else:
                rows, col_mapping = self._parse_pdf(content)
        except Exception as e:
            elapsed = (time.time() - start_time) * 1000
            return ParsedIngestionResult(
                source_file=source_file,
                entry_type=self.entry_type,
                entries=[],
                row_count=0,
                error_count=1,
                errors=[{"row": 0, "error": f"Failed to parse census report: {e}"}],
                column_mapping={},
                processing_time_ms=round(elapsed, 2),
            )

        entries: list[GrossGLTransaction] = []
        for row in rows:
            try:
                entry = GrossGLTransaction(
                    account_number=sanitize_string(row.get("department_code", row.get("unit", ""))),
                    account_description=sanitize_string(row.get("description", row.get("department_name", ""))),
                    debit_amount=sanitize_amount(row.get("patient_days", row.get("visits", "0"))),
                    credit_amount=Decimal("0.00"),
                    period=sanitize_string(row.get("period", row.get("month", ""))),
                    department_code=sanitize_string(row.get("department_code", row.get("unit", ""))),
                    department_name=sanitize_string(row.get("department_name", row.get("description", ""))),
                    source_file=source_file,
                    row_index=row.get("_row_index"),
                )
                entries.append(entry)
            except Exception as e:
                errors.append({"row": row.get("_row_index"), "error": str(e)})

        elapsed = (time.time() - start_time) * 1000
        return ParsedIngestionResult(
            source_file=source_file,
            entry_type=self.entry_type,
            entries=entries,
            row_count=len(entries),
            error_count=len(errors),
            errors=errors,
            column_mapping=col_mapping,
            processing_time_ms=round(elapsed, 2),
        )

    def _parse_pdf(self, content: bytes) -> tuple[list[dict], dict[str, str]]:
        """Extract census data from PDFs using pdfplumber."""
        import pdfplumber

        rows: list[dict] = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if not table or len(table) < 2:
                            continue
                        headers = [str(c or "").strip() for c in table[0]]
                        col_mapping = detect_column_mapping(headers)
                        for i, row_data in enumerate(table[1:], start=2):
                            row_dict: dict = {"_row_index": i, "source": "pdf_table"}
                            for j, cell in enumerate(row_data):
                                if j < len(headers):
                                    row_dict[headers[j].lower().replace(" ", "_")] = cell
                            rows.append(row_dict)
                else:
                    # Fallback: parse text lines
                    lines = text.split("\n")
                    for i, line in enumerate(lines):
                        parts = re.split(r"\s{2,}", line.strip())
                        if len(parts) >= 3:
                            rows.append({
                                "_row_index": i + 1,
                                "source": "pdf_text",
                                "department": parts[0],
                                "patient_days": parts[-1] if parts[-1].replace(".", "").isdigit() else "0",
                            })
        return rows, {}

    def _parse_excel(self, content: bytes, ext: str) -> tuple[list[dict], dict[str, str]]:
        """Parse Excel census reports."""
        import openpyxl

        rows: list[dict] = []
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        if sheet.max_row and sheet.max_row >= 2:
            headers = [str(cell.value or "") for cell in sheet[1]]
            col_mapping = detect_column_mapping(headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict: dict = {"_row_index": row_idx}
                for col_idx, header in enumerate(headers):
                    val = row[col_idx] if col_idx < len(row) else None
                    row_dict[header.lower().replace(" ", "_")] = val
                rows.append(row_dict)
        workbook.close()
        return rows, {}

    def _parse_csv_or_tsv(self, content: bytes, ext: str) -> tuple[list[dict], dict[str, str]]:
        """Parse CSV/TSV census data."""
        import csv
        import io

        delimiter = "\t" if ext == ".tsv" else ","
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows: list[dict] = []

        try:
            headers = next(reader)
        except StopIteration:
            return rows, {}

        headers = [h.strip() for h in headers]
        for row_idx, row in enumerate(reader, start=2):
            row_dict: dict = {"_row_index": row_idx}
            for col_idx, val in enumerate(row):
                if col_idx < len(headers):
                    row_dict[headers[col_idx].lower().replace(" ", "_")] = val.strip()
            rows.append(row_dict)

        return rows, {}
