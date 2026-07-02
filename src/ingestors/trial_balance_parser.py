"""
Trial Balance Parser — handles Excel (.xlsx, .xls), CSV, and TSV files.

Auto-detects column mappings from headers, sanitizes currency values,
handles unbalanced entries, and produces structured GrossGLTransaction records.
"""

from __future__ import annotations

import csv
import io
import time
from pathlib import Path
from typing import Any, Optional

import openpyxl
import xlrd

from src.core.models import EntryType, GrossGLTransaction, ParsedIngestionResult
from src.ingestors.base_parser import BaseParser, detect_column_mapping, sanitize_amount, sanitize_string


class TrialBalanceParser(BaseParser):
    """Parser for trial balance files in Excel (xlsx/xls), CSV, and TSV formats."""

    entry_type = EntryType.TRIAL_BALANCE

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv", ".txt"}

    def parse(self, content: bytes, source_file: str, **kwargs) -> ParsedIngestionResult:
        """Parse a trial balance file. Auto-detects format by extension."""
        start_time = time.time()
        ext = Path(source_file).suffix.lower()

        if ext == ".xlsx":
            rows, col_mapping, errors = self._parse_xlsx(content, source_file)
        elif ext == ".xls":
            rows, col_mapping, errors = self._parse_xls(content, source_file)
        elif ext in (".csv", ".txt"):
            rows, col_mapping, errors = self._parse_csv(content, source_file, delimiter=",")
        elif ext == ".tsv":
            rows, col_mapping, errors = self._parse_csv(content, source_file, delimiter="\t")
        else:
            # Try CSV first, fall back to XLSX
            try:
                rows, col_mapping, errors = self._parse_csv(content, source_file, delimiter=",")
            except Exception:
                rows, col_mapping, errors = self._parse_xlsx(content, source_file)

        elapsed = (time.time() - start_time) * 1000

        entries: list[GrossGLTransaction] = []
        for row in rows:
            try:
                entry = GrossGLTransaction(
                    account_number=sanitize_string(row.get("account_number", "")),
                    account_description=sanitize_string(row.get("account_description", "")),
                    debit_amount=sanitize_amount(row.get("debit_amount", "0")),
                    credit_amount=sanitize_amount(row.get("credit_amount", "0")),
                    period=sanitize_string(row.get("period", "")),
                    department_code=sanitize_string(row.get("department_code", "")),
                    department_name=sanitize_string(row.get("department_name", "")),
                    source_file=source_file,
                    row_index=row.get("_row_index"),
                )
                entries.append(entry)
            except Exception as e:
                errors.append({"row": row.get("_row_index"), "error": str(e)})

        return ParsedIngestionResult(
            source_file=source_file,
            entry_type=self.entry_type,
            entries=entries,
            row_count=len(entries),
            error_count=len(errors),
            errors=errors,
            column_mapping=col_mapping,
            processing_time_ms=round(elapsed, 2),
        )

    def _parse_xlsx(
        self, content: bytes, source_file: str
    ) -> tuple[list[dict], dict[str, str], list[dict]]:
        """Parse .xlsx files using openpyxl."""
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows: list[dict] = []
        errors: list[dict] = []

        # Sheet might be empty
        if sheet.max_row is None or sheet.max_row < 2:
            return rows, {}, errors

        # Extract headers from first row
        headers = [str(cell.value or "") for cell in sheet[1]]
        col_mapping = detect_column_mapping(headers)

        # Build header index
        header_index: dict[str, int] = {}
        for field, header in col_mapping.items():
            try:
                header_index[field] = headers.index(header)
            except ValueError:
                pass

        # Parse data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict: dict = {"_row_index": row_idx}
            for field, col_idx in header_index.items():
                val = row[col_idx] if col_idx < len(row) else None
                row_dict[field] = val
            rows.append(row_dict)

        workbook.close()
        return rows, col_mapping, errors

    def _parse_xls(
        self, content: bytes, source_file: str
    ) -> tuple[list[dict], dict[str, str], list[dict]]:
        """Parse legacy .xls files using xlrd."""
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        rows: list[dict] = []
        errors: list[dict] = []

        if sheet.nrows < 2:
            return rows, {}, errors

        headers = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        col_mapping = detect_column_mapping(headers)

        header_index: dict[str, int] = {}
        for field, header in col_mapping.items():
            try:
                header_index[field] = headers.index(header)
            except ValueError:
                pass

        for row_idx in range(1, sheet.nrows):
            row_dict: dict = {"_row_index": row_idx + 1}
            for field, col_idx in header_index.items():
                val = sheet.cell_value(row_idx, col_idx)
                row_dict[field] = val
            rows.append(row_dict)

        return rows, col_mapping, errors

    def _parse_csv(
        self, content: bytes, source_file: str, delimiter: str = ","
    ) -> tuple[list[dict], dict[str, str], list[dict]]:
        """Parse CSV/TSV files."""
        text = content.decode("utf-8-sig")  # Handles BOM
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows: list[dict] = []
        errors: list[dict] = []

        try:
            headers = next(reader)
        except StopIteration:
            return rows, {}, errors

        headers = [h.strip() for h in headers]
        col_mapping = detect_column_mapping(headers)

        # Determine which columns to extract
        extract_fields = list(col_mapping.keys())
        header_positions: dict[str, int] = {}
        for field, header in col_mapping.items():
            try:
                header_positions[field] = headers.index(header)
            except ValueError:
                pass

        for row_idx, row in enumerate(reader, start=2):
            row_dict: dict = {"_row_index": row_idx}
            for field in extract_fields:
                col_idx = header_positions.get(field)
                if col_idx is not None and col_idx < len(row):
                    row_dict[field] = row[col_idx].strip()
            rows.append(row_dict)

        return rows, col_mapping, errors
